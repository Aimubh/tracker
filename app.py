"""
Flask app — Lazer Believe Ranking Tracker
Source of truth: product_catalog.json (built from Active Listings Links.xlsx)
Run: python app.py  →  open http://localhost:5000
"""

import asyncio
import json
import os
import sys

# Force UTF-8 stdout so scraper print() statements containing ₹ and other
# non-ASCII characters don't crash on Windows' default cp1252 console encoding.
# (A failed print inside a scraper's try-block would otherwise discard valid
# scraped data and surface as a misleading "could not scrape" error.)
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

from flask import Flask, render_template, request, jsonify

app = Flask(__name__)

AMAZON_CATALOG_PATH  = os.path.join(os.path.dirname(__file__), "amazon_catalog.json")
BLINKIT_CATALOG_PATH = os.path.join(os.path.dirname(__file__), "blinkit_catalog.json")


# ── Catalog helpers ────────────────────────────────────────────────────────────

def load_amazon() -> list:
    with open(AMAZON_CATALOG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def load_blinkit() -> list:
    with open(BLINKIT_CATALOG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def save_amazon(data: list):
    with open(AMAZON_CATALOG_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def save_blinkit(data: list):
    with open(BLINKIT_CATALOG_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


# ── Routes ─────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("dashboard.html")


@app.route("/api/products", methods=["GET"])
def api_products():
    """Returns all products from both catalogs for the dropdown."""
    try:
        amazon_entries  = load_amazon()
        blinkit_entries = load_blinkit()

        amazon_products = [
            {
                "url":          e["url"],
                "title":        e["title"],
                "asin":         e["asin"],
                "price":        e.get("price", 0),
                "rating":       e.get("rating", 0),
                "review_count": e.get("review_count", 0),
                "image_count":  e.get("image_count", 0),
                "scraped":      e.get("scraped", False),
                "platform":     "amazon",
            }
            for e in amazon_entries
        ]

        blinkit_products = [
            {
                "url":          e["url"],
                "title":        e["title"],
                "product_id":   e["product_id"],
                "price":        e.get("price", 0),
                "image_count":  e.get("image_count", 0),
                "scraped":      e.get("scraped", False),
                "platform":     "blinkit",
            }
            for e in blinkit_entries
        ]

        unscraped_amazon  = sum(1 for e in amazon_entries  if not e.get("scraped"))
        unscraped_blinkit = sum(1 for e in blinkit_entries if not e.get("scraped"))

        return jsonify({
            "amazon":            amazon_products,
            "blinkit":           blinkit_products,
            "total":             len(amazon_products) + len(blinkit_products),
            "unscraped_amazon":  unscraped_amazon,
            "unscraped_blinkit": unscraped_blinkit,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/analyze", methods=["POST"])
def api_analyze():
    """
    Accepts a product URL (Amazon or Blinkit) from our catalog.
    Scrapes that listing live, then finds top competitors via image search,
    and returns gap analysis.
    """
    data = request.get_json()
    product_url = data.get("product_url", "").strip()
    platform    = data.get("platform", "").strip()   # "amazon" or "blinkit"

    if not product_url or not platform:
        return jsonify({"error": "product_url and platform are required"}), 400

    try:
        from scrapers.amazon import scrape_our_amazon_listing, scrape_amazon_by_image, scrape_amazon
        from scrapers.blinkit import scrape_our_blinkit_listing, scrape_blinkit
        from analyzer import analyze

        our_data = None
        competitors = []
        warnings = []

        if platform == "amazon":
            # Step 1: use cached catalog data if already scraped, otherwise scrape live
            catalog_entry = _get_catalog_entry(product_url, "amazon")
            if catalog_entry and catalog_entry.get("scraped") and catalog_entry.get("title") and catalog_entry["title"] != catalog_entry.get("asin"):
                print(f"[App] Using cached Amazon data for {catalog_entry['title'][:60]}")
                our_data = {**catalog_entry, "source": "amazon_ours", "platform": "amazon",
                            "description_word_count": len((catalog_entry.get("description") or "").split()),
                            "bullet_count": catalog_entry.get("bullet_count", 0),
                            "has_best_seller_badge": False}
            else:
                print(f"[App] Live-scraping Amazon listing: {product_url}")
                our_data = asyncio.run(scrape_our_amazon_listing(product_url))
                if our_data and our_data.get("title"):
                    _update_catalog_entry(product_url, "amazon", our_data)

            if not our_data or not our_data.get("title"):
                return jsonify({"error": "Could not load your Amazon listing. Try running scrape_catalog.py first."}), 400

            # Step 2: image search for competitors
            image_url = our_data.get("main_image_url", "")
            if image_url:
                print(f"[App] Using image search: {image_url[:80]}")
                competitors = asyncio.run(scrape_amazon_by_image(image_url, max_results=5))
                if not competitors:
                    warnings.append("Image search returned no results — falling back to keyword search.")

            # Step 3: fallback to keyword search
            if not competitors:
                keyword = _make_keyword(our_data.get("title", ""))
                competitors = asyncio.run(scrape_amazon(keyword, max_results=5))
                if not competitors:
                    return jsonify({"error": f"No competitors found on Amazon for '{keyword}'."}), 400

        elif platform == "blinkit":
            # Step 1: use cached catalog data if already scraped, otherwise scrape live
            catalog_entry = _get_catalog_entry(product_url, "blinkit")
            if catalog_entry and catalog_entry.get("scraped") and catalog_entry.get("title"):
                print(f"[App] Using cached Blinkit data for {catalog_entry['title'][:60]}")
                our_data = {**catalog_entry, "source": "blinkit_ours", "platform": "blinkit",
                            "description_word_count": len((catalog_entry.get("description") or "").split()),
                            "bullet_count": catalog_entry.get("bullet_count", 0),
                            "has_best_seller_badge": False}
            else:
                print(f"[App] Live-scraping Blinkit listing: {product_url}")
                our_data = asyncio.run(scrape_our_blinkit_listing(product_url))
                if our_data and our_data.get("title"):
                    _update_catalog_entry(product_url, "blinkit", our_data)

            if not our_data or not our_data.get("title"):
                return jsonify({"error": "Could not load your Blinkit listing."}), 400

            # Step 2: keyword search on Blinkit for competitors
            keyword = _make_keyword(our_data.get("title", ""))
            competitors = asyncio.run(scrape_blinkit(keyword, max_results=5))
            if not competitors:
                warnings.append(f"Blinkit returned no results for '{keyword}'.")
                return jsonify({"error": f"No competitors found on Blinkit for '{keyword}'.", "warnings": warnings}), 400

        else:
            return jsonify({"error": "platform must be 'amazon' or 'blinkit'"}), 400

        # Filter out own-brand, self, and off-category competitors before scoring
        competitors, filter_warnings = _filter_competitors(our_data, competitors)
        warnings.extend(filter_warnings)
        if not competitors:
            return jsonify({"error": "No relevant competitors left after filtering.",
                            "warnings": warnings}), 400

        # Benchmark against the review leader: the competitor with the most reviews
        # becomes #1 everywhere (gap table, image comparison, Roadmap to #1).
        # Blinkit results carry no review data, so only re-rank Amazon.
        if platform == "amazon" and any(c.get("review_count", 0) for c in competitors):
            competitors.sort(key=lambda c: c.get("review_count", 0) or 0, reverse=True)
            warnings.append(
                f"Benchmarking against the most-reviewed competitor "
                f"({competitors[0].get('review_count', 0):,} reviews)."
            )

        result = analyze(our_data, competitors)
        result["warnings"]    = warnings
        result["platform"]    = platform
        result["our_product"] = our_data
        return jsonify(result)

    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/api/scrape-status", methods=["GET"])
def api_scrape_status():
    """Returns how many entries are still unscraped — for dashboard progress bar."""
    try:
        amazon_entries  = load_amazon()
        blinkit_entries = load_blinkit()
        return jsonify({
            "amazon":  {"total": len(amazon_entries),  "scraped": sum(1 for e in amazon_entries  if e.get("scraped"))},
            "blinkit": {"total": len(blinkit_entries), "scraped": sum(1 for e in blinkit_entries if e.get("scraped"))},
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/add-product", methods=["POST"])
def api_add_product():
    """
    Accepts a raw Amazon or Blinkit product URL, scrapes it live, and adds it to
    the catalog so it can be analyzed immediately.
    Body: {"url": "https://www.amazon.in/dp/XXXX"}
    """
    import re
    data = request.get_json() or {}
    url = (data.get("url") or "").strip()

    if not url or not url.startswith("http"):
        return jsonify({"error": "Please provide a full product URL starting with http(s)://"}), 400

    # Detect platform
    if "amazon." in url:
        platform = "amazon"
    elif "blinkit.com" in url:
        platform = "blinkit"
    else:
        return jsonify({"error": "URL must be an amazon.in or blinkit.com product link."}), 400

    try:
        if platform == "amazon":
            # Extract ASIN from /dp/XXXX or /gp/product/XXXX
            m = re.search(r"/(?:dp|gp/product|product)/([A-Z0-9]{10})", url, re.I)
            if not m:
                return jsonify({"error": "Could not find an ASIN in that Amazon URL (expected /dp/ASIN)."}), 400
            asin = m.group(1).upper()
            clean_url = f"https://www.amazon.in/dp/{asin}"

            entries = load_amazon()
            existing = next((e for e in entries if e.get("asin") == asin), None)
            if existing and existing.get("scraped"):
                return jsonify({"ok": True, "already_exists": True, "platform": "amazon",
                                "product": _amazon_public(existing),
                                "message": "This product is already in your catalog."})

            from scrapers.amazon import scrape_our_amazon_listing
            scraped = _scrape_with_retry(scrape_our_amazon_listing, clean_url)
            if not scraped or not scraped.get("title"):
                return jsonify({"error": "Could not scrape that Amazon listing after retrying — "
                                         "Amazon may be showing a bot-check. Please try again in a moment."}), 400

            entry = existing or {"asin": asin, "url": clean_url}
            entry.update({
                "asin":           asin,
                "url":            clean_url,
                "title":          scraped.get("title", asin),
                "price":          scraped.get("price", 0),
                "rating":         scraped.get("rating", 0),
                "review_count":   scraped.get("review_count", 0),
                "image_count":    scraped.get("image_count", 0),
                "image_urls":     scraped.get("image_urls", []),
                "main_image_url": scraped.get("main_image_url", ""),
                "description":    scraped.get("description", ""),
                "bullet_count":   scraped.get("bullet_count", 0),
                "keywords":       scraped.get("keywords", []),
                "scraped":        True,
            })
            if not existing:
                entries.append(entry)
            save_amazon(entries)
            return jsonify({"ok": True, "platform": "amazon", "product": _amazon_public(entry)})

        else:  # blinkit
            m = re.search(r"/prid/(\d+)", url) or re.search(r"product_id=(\d+)", url)
            if not m:
                return jsonify({"error": "Could not find a product ID in that Blinkit URL (expected /prid/NUMBER)."}), 400
            pid = m.group(1)
            clean_url = f"https://blinkit.com/prn/prid/{pid}"

            entries = load_blinkit()
            existing = next((e for e in entries if e.get("product_id") == pid), None)
            if existing and existing.get("scraped"):
                return jsonify({"ok": True, "already_exists": True, "platform": "blinkit",
                                "product": _blinkit_public(existing),
                                "message": "This product is already in your catalog."})

            from scrapers.blinkit import scrape_our_blinkit_listing
            scraped = _scrape_with_retry(scrape_our_blinkit_listing, clean_url)
            if not scraped or not scraped.get("title"):
                return jsonify({"error": "Could not scrape that Blinkit listing after retrying — "
                                         "it may be unavailable in the Mumbai region. Please try again."}), 400

            entry = existing or {"product_id": pid, "url": clean_url, "slug": "added-manually"}
            entry.update({
                "product_id":     pid,
                "url":            clean_url,
                "title":          scraped.get("title", f"Product #{pid}"),
                "price":          scraped.get("price", 0),
                "image_count":    scraped.get("image_count", 0),
                "image_urls":     scraped.get("image_urls", []),
                "main_image_url": scraped.get("main_image_url", ""),
                "description":    scraped.get("description", ""),
                "keywords":       scraped.get("keywords", []),
                "scraped":        True,
            })
            if not existing:
                entries.append(entry)
            save_blinkit(entries)
            return jsonify({"ok": True, "platform": "blinkit", "product": _blinkit_public(entry)})

    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


EXCEL_PATH = os.path.join(os.path.expanduser("~"), "Downloads", "Active Listings Links.xlsx")


def _merge_excel_workbook(wb) -> dict:
    """
    Merge an openpyxl workbook into the catalogs: append any ASIN / product-ID
    not already present as an unscraped placeholder. Existing entries (and their
    scraped data) are left untouched — this is a merge, not an overwrite.
    Returns a summary dict.
    """
    # ── Amazon ──────────────────────────────────────────────────────────────
    amazon = load_amazon()
    have_asins = {e.get("asin") for e in amazon}
    new_amazon = 0
    if "Amazon" in wb.sheetnames:
        seen = set()
        for row in wb["Amazon"].iter_rows(min_row=2, values_only=True):
            asin = str(row[0]).strip() if row[0] else ""
            base = str(row[1]).strip() if row[1] else ""
            if not asin or not base or not base.startswith("http") or asin in seen:
                continue
            seen.add(asin)
            if asin in have_asins:
                continue
            amazon.append({
                "asin": asin, "url": f"https://www.amazon.in/dp/{asin}",
                "title": asin, "price": 0.0, "rating": 0.0, "review_count": 0,
                "image_count": 0, "image_urls": [], "main_image_url": "",
                "description": "", "bullet_count": 0, "keywords": [], "scraped": False,
            })
            have_asins.add(asin)
            new_amazon += 1

    # ── Blinkit ─────────────────────────────────────────────────────────────
    blinkit = load_blinkit()
    have_pids = {e.get("product_id") for e in blinkit}
    new_blinkit = 0
    if "Blinkit" in wb.sheetnames:
        seen = set()
        for row in wb["Blinkit"].iter_rows(min_row=2, values_only=True):
            pid = row[0]
            base = str(row[1]).strip() if row[1] else ""
            if not pid or not base or not base.startswith("http"):
                continue
            try:
                pid_str = str(int(pid))
            except (ValueError, TypeError):
                pid_str = str(pid).strip()
            if not pid_str or pid_str in seen or pid_str in have_pids:
                continue
            seen.add(pid_str)
            blinkit.append({
                "product_id": pid_str, "url": f"https://blinkit.com/prn/prid/{pid_str}",
                "slug": "synced-from-excel", "title": f"Product #{pid_str}",
                "price": 0.0, "rating": 0.0, "review_count": 0, "image_count": 0,
                "image_urls": [], "main_image_url": "", "description": "",
                "bullet_count": 0, "keywords": [], "scraped": False,
            })
            have_pids.add(pid_str)
            new_blinkit += 1

    if new_amazon:
        save_amazon(amazon)
    if new_blinkit:
        save_blinkit(blinkit)

    return {
        "ok": True,
        "new_amazon": new_amazon, "new_blinkit": new_blinkit,
        "pending_amazon":  sum(1 for e in amazon  if not e.get("scraped")),
        "pending_blinkit": sum(1 for e in blinkit if not e.get("scraped")),
        "total_amazon": len(amazon), "total_blinkit": len(blinkit),
    }


@app.route("/api/sync-excel", methods=["POST"])
def api_sync_excel():
    """Sync from the fixed Downloads path (quick shortcut when the file is there)."""
    if not os.path.exists(EXCEL_PATH):
        return jsonify({"ok": False,
                        "error": f"No Excel found at {EXCEL_PATH}. Use 'Upload Excel' to pick a file instead."}), 400
    try:
        import openpyxl
    except ImportError:
        return jsonify({"ok": False, "error": "openpyxl not installed. Run: pip install openpyxl"}), 400
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
        result = _merge_excel_workbook(wb)
        wb.close()
        return jsonify(result)
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/api/upload-excel", methods=["POST"])
def api_upload_excel():
    """
    Accepts an uploaded .xlsx (multipart form field 'file'), merges it into the
    catalogs. Same merge semantics as sync — adds only new products.
    """
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"ok": False, "error": "No file uploaded."}), 400
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"ok": False, "error": "Please upload an .xlsx file."}), 400

    try:
        import openpyxl
    except ImportError:
        return jsonify({"ok": False, "error": "openpyxl not installed. Run: pip install openpyxl"}), 400

    try:
        import io
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), read_only=True)
        if "Amazon" not in wb.sheetnames and "Blinkit" not in wb.sheetnames:
            wb.close()
            return jsonify({"ok": False,
                            "error": "Workbook has no 'Amazon' or 'Blinkit' sheet. "
                                     "Expected the Active Listings Links format."}), 400
        result = _merge_excel_workbook(wb)
        wb.close()
        result["filename"] = file.filename
        return jsonify(result)
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/api/scrape-batch", methods=["POST"])
def api_scrape_batch():
    """
    Scrapes a small batch of unscraped entries for one platform and saves them.
    Called repeatedly by the UI so progress can be shown without one giant
    request. Body: {"platform": "amazon"|"blinkit", "batch_size": 3}
    Returns what was scraped this call plus how many remain.
    """
    data = request.get_json() or {}
    platform   = data.get("platform", "amazon")
    batch_size = max(1, min(int(data.get("batch_size", 3)), 10))

    if platform not in ("amazon", "blinkit"):
        return jsonify({"ok": False, "error": "platform must be 'amazon' or 'blinkit'"}), 400

    try:
        entries = load_amazon() if platform == "amazon" else load_blinkit()
        pending = [e for e in entries if not e.get("scraped")]
        if not pending:
            return jsonify({"ok": True, "done": True, "scraped_now": [], "remaining": 0})

        if platform == "amazon":
            from scrapers.amazon import scrape_our_amazon_listing
            scrape_fn = scrape_our_amazon_listing
        else:
            from scrapers.blinkit import scrape_our_blinkit_listing
            scrape_fn = scrape_our_blinkit_listing

        scraped_now = []
        for entry in pending[:batch_size]:
            res = _scrape_with_retry(scrape_fn, entry["url"])
            if res and res.get("title"):
                entry.update({
                    "title":          res.get("title", entry.get("title")),
                    "price":          res.get("price", 0),
                    "rating":         res.get("rating", entry.get("rating", 0)),
                    "review_count":   res.get("review_count", entry.get("review_count", 0)),
                    "image_count":    res.get("image_count", 0),
                    "image_urls":     res.get("image_urls", []),
                    "main_image_url": res.get("main_image_url", ""),
                    "description":    res.get("description", ""),
                    "bullet_count":   res.get("bullet_count", entry.get("bullet_count", 0)),
                    "keywords":       res.get("keywords", []),
                    "scraped":        True,
                })
                scraped_now.append({"title": entry["title"][:60], "ok": True})
            else:
                # Mark as scraped so we don't loop forever on a dead listing
                entry["scraped"] = True
                scraped_now.append({"title": entry.get("title", "?")[:60], "ok": False})

        if platform == "amazon":
            save_amazon(entries)
        else:
            save_blinkit(entries)

        remaining = sum(1 for e in entries if not e.get("scraped"))
        return jsonify({"ok": True, "done": remaining == 0,
                        "scraped_now": scraped_now, "remaining": remaining})
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/api/ai-rewrite", methods=["POST"])
def api_ai_rewrite():
    """
    Uses Claude to rewrite our listing (title, bullets, strategy) to outrank
    competitors. Requires ANTHROPIC_API_KEY; degrades gracefully if absent.
    Body: {"our_product": {...}, "competitors": [...], "platform": "amazon"}
    """
    if not os.environ.get("ANTHROPIC_API_KEY"):
        return jsonify({
            "ok": False,
            "needs_key": True,
            "error": "AI rewrite needs a Claude API key. Set the ANTHROPIC_API_KEY "
                     "environment variable and restart the server to enable it.",
        }), 400

    data = request.get_json() or {}
    our = data.get("our_product") or {}
    competitors = data.get("competitors") or []
    platform = data.get("platform", "amazon")

    if not our.get("title"):
        return jsonify({"ok": False, "error": "Missing our_product data."}), 400

    try:
        import anthropic
    except ImportError:
        return jsonify({
            "ok": False,
            "error": "The 'anthropic' package isn't installed. Run: pip install anthropic",
        }), 400

    # Build a compact competitor summary for the prompt
    comp_lines = []
    for i, c in enumerate(competitors[:5], 1):
        comp_lines.append(
            f"{i}. {c.get('title','')[:120]} | ₹{c.get('price',0)} | "
            f"{c.get('rating',0)}★ | {c.get('review_count',0)} reviews | "
            f"{c.get('image_count',0)} images"
        )
    comp_block = "\n".join(comp_lines) or "(no competitor data)"

    our_block = (
        f"Title: {our.get('title','')}\n"
        f"Price: ₹{our.get('price',0)}\n"
        f"Rating: {our.get('rating',0)}★ | Reviews: {our.get('review_count',0)}\n"
        f"Images: {our.get('image_count',0)} | Bullets: {our.get('bullet_count',0)}\n"
        f"Current bullets/description:\n{(our.get('description') or '(none)')[:1500]}"
    )

    SCHEMA = {
        "type": "object",
        "properties": {
            "optimized_title": {"type": "string"},
            "bullets": {"type": "array", "items": {"type": "string"}},
            "strategy": {"type": "string"},
            "backend_keywords": {"type": "array", "items": {"type": "string"}},
        },
        "required": ["optimized_title", "bullets", "strategy", "backend_keywords"],
        "additionalProperties": False,
    }

    prompt = (
        f"You are an expert {platform.title()} India listing strategist. "
        f"Rewrite our product listing so it outranks the competitors below.\n\n"
        f"=== OUR LISTING ===\n{our_block}\n\n"
        f"=== TOP COMPETITORS ===\n{comp_block}\n\n"
        f"Produce:\n"
        f"1. optimized_title — a keyword-rich {platform.title()} title (<=200 chars), brand 'Lazer' first, "
        f"front-loading the highest-traffic search terms, no keyword stuffing.\n"
        f"2. bullets — 5 benefit-led bullet points (each <=220 chars) covering material, "
        f"dimensions/capacity, key use-cases, what's included, and care/quality.\n"
        f"3. strategy — 2-3 sentences on the single biggest lever to outrank the #1 competitor.\n"
        f"4. backend_keywords — 10-15 search terms NOT already in the title, for the backend field.\n"
        f"Write for Indian shoppers. Be specific to THIS product, not generic."
    )

    try:
        client = anthropic.Anthropic()
        msg = client.messages.create(
            model="claude-opus-4-8",
            max_tokens=4000,
            thinking={"type": "adaptive"},
            output_config={"format": {"type": "json_schema", "schema": SCHEMA}},
            messages=[{"role": "user", "content": prompt}],
        )
        text = next((b.text for b in msg.content if b.type == "text"), "")
        result = json.loads(text)
        return jsonify({"ok": True, "rewrite": result})
    except anthropic.AuthenticationError:
        return jsonify({"ok": False, "needs_key": True,
                        "error": "Invalid ANTHROPIC_API_KEY."}), 400
    except Exception as e:
        return jsonify({"ok": False, "error": f"AI rewrite failed: {e}"}), 500


# ── Helpers ────────────────────────────────────────────────────────────────────

def _scrape_with_retry(scrape_fn, url: str, attempts: int = 2):
    """
    Run an async scrape function with a couple of retries — Amazon/Blinkit
    occasionally serve a bot-check or slow page on a cold request, and a single
    transient miss shouldn't block adding a valid product.
    """
    last = None
    for i in range(attempts):
        try:
            result = asyncio.run(scrape_fn(url))
            if result and result.get("title"):
                return result
            last = result
        except Exception as e:
            print(f"[AddProduct] scrape attempt {i+1} failed: {e}")
    return last


def _amazon_public(e: dict) -> dict:
    return {"url": e["url"], "title": e["title"], "asin": e["asin"],
            "price": e.get("price", 0), "rating": e.get("rating", 0),
            "review_count": e.get("review_count", 0), "image_count": e.get("image_count", 0),
            "scraped": e.get("scraped", False), "platform": "amazon"}


def _blinkit_public(e: dict) -> dict:
    return {"url": e["url"], "title": e["title"], "product_id": e["product_id"],
            "price": e.get("price", 0), "image_count": e.get("image_count", 0),
            "scraped": e.get("scraped", False), "platform": "blinkit"}

def _make_keyword(title: str) -> str:
    import re
    stopwords = {"lazer", "pure", "mini", "new", "the", "a", "an", "and", "in",
                 "for", "with", "of", "1", "set", "ml", "inch", "kit", "pcs",
                 "pack", "cm", "by", "from"}
    words = re.findall(r"\b[a-zA-Z0-9]+\b", title.lower())
    filtered = [w for w in words if w not in stopwords]
    return " ".join(filtered[:5])


# Product-type signatures. Each type lists keywords that, if present in a title,
# mark the product as belonging to that type. Used to keep competitors in the same
# category as our product (e.g. don't compare a steel grooming kit to a face cream).
_PRODUCT_TYPES = {
    "cosmetic":   ["cream", "gel", "lotion", "serum", "mask", "scrub", "bomb",
                   "moistur", "lightening", "softening", "spa", "wax", "balm",
                   "oil", "shampoo", "soap", "facial", "bleach"],
    "book":       ["book", "coloring book", "colouring book", "activity book",
                   "notebook", "diary", "journal", "sketchbook"],
    "tool_kit":   ["clipper", "cutter", "scissor", "tweezer", "grooming",
                   "stainless steel", "manicure set", "pedicure set", "tool",
                   "nail file", "trimmer", "kit with"],
    "bottle":     ["bottle", "dispenser", "container", "jar", "flask",
                   "toiletries", "refillable", "pump"],
    "sewing":     ["sewing", "needle", "thread", "stitch"],
    "stationery": ["pencil", "pen", "marker", "crayon", "eraser", "sharpener",
                   "pouch", "doodle", "canvas bag"],
}


def _classify_type(title: str) -> set:
    """Return the set of product-types whose keywords appear in the title."""
    t = (title or "").lower()
    types = {name for name, kws in _PRODUCT_TYPES.items()
             if any(kw in t for kw in kws)}
    return types


def _filter_competitors(our_data: dict, competitors: list) -> tuple[list, list]:
    """
    Remove competitors that are our own brand, the exact same listing, or a
    different product type than ours. Falls back to the unfiltered list if
    filtering would leave too few to be useful. Returns (filtered, warnings).
    """
    warnings = []
    if not competitors:
        return competitors, warnings

    our_title = (our_data.get("title") or "")
    our_url   = our_data.get("url", "")
    our_types = _classify_type(our_title)

    kept, dropped_brand, dropped_self, dropped_type = [], 0, 0, 0
    for c in competitors:
        c_title = (c.get("title") or "")
        c_url   = c.get("url", "")

        # Exclude our own brand
        if "lazer" in c_title.lower():
            dropped_brand += 1
            continue
        # Exclude the exact same listing (self-comparison)
        if c_url and our_url and c_url == our_url:
            dropped_self += 1
            continue
        # Exclude obvious product-type mismatches (only when we could classify ours)
        if our_types:
            c_types = _classify_type(c_title)
            # Keep if the competitor shares any of our types, or if we can't tell
            # what type it is (don't over-filter on ambiguous titles).
            if c_types and not (c_types & our_types):
                dropped_type += 1
                continue
        kept.append(c)

    if dropped_brand:
        warnings.append(f"Excluded {dropped_brand} of your own listings from competitors.")
    if dropped_type:
        warnings.append(f"Excluded {dropped_type} off-category result(s) (different product type).")

    # Don't return a uselessly small set — need at least 2 to benchmark against.
    if len(kept) < 2:
        warnings.append("Too few same-category competitors found — showing all results instead.")
        return competitors, warnings

    return kept, warnings


def _get_catalog_entry(url: str, platform: str) -> dict | None:
    """Find a catalog entry by URL."""
    entries = load_amazon() if platform == "amazon" else load_blinkit()
    for entry in entries:
        if entry["url"] == url:
            return entry
    return None


def _update_catalog_entry(url: str, platform: str, data: dict):
    """Update a catalog entry with freshly scraped data and save to the correct db file."""
    try:
        entries = load_amazon() if platform == "amazon" else load_blinkit()
        for entry in entries:
            if entry["url"] == url:
                entry.update({
                    "title":          data.get("title", entry["title"]),
                    "price":          data.get("price", entry.get("price", 0)),
                    "rating":         data.get("rating", entry.get("rating", 0)),
                    "review_count":   data.get("review_count", entry.get("review_count", 0)),
                    "image_count":    data.get("image_count", entry.get("image_count", 0)),
                    "image_urls":     data.get("image_urls", entry.get("image_urls", [])),
                    "main_image_url": data.get("main_image_url", entry.get("main_image_url", "")),
                    "description":    data.get("description", entry.get("description", "")),
                    "bullet_count":   data.get("bullet_count", entry.get("bullet_count", 0)),
                    "keywords":       data.get("keywords", entry.get("keywords", [])),
                    "scraped":        True,
                })
                break
        if platform == "amazon":
            save_amazon(entries)
        else:
            save_blinkit(entries)
    except Exception as e:
        print(f"[App] Could not update catalog: {e}")


if __name__ == "__main__":
    print("Starting Lazer Believe Ranking Tracker...")
    print("Open http://localhost:5000")
    app.run(debug=False, port=5000)
