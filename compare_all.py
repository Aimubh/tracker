"""
Builds a side-by-side comparison Excel: each of our Amazon products vs its top
competitor (the most-reviewed, SAME-PRODUCT result — not just same category).

Run:  python compare_all.py
- Resumable: progress is saved to compare_progress.json after every product, so
  an interruption (CAPTCHA, timeout, Ctrl-C) is safe — just re-run to continue.
- Output: product_comparison.xlsx

What changed vs the old version (which matched wrong products):
- STRICTER matching: a competitor must share our product's *core noun* (the most
  specific product word, e.g. "manicure", "saree", "fryer"), not merely the same
  loose category. Broken/junk rows (single-word brand titles, no price, no rating)
  are dropped.
- RICHER comparison: price delta + who's cheaper, rating gap, reviews gap,
  missing keywords (what the competitor ranks for that our listing lacks),
  bullet-count gap, image counts, and EMBEDDED thumbnails for both products.
- Garbled text from bad-encoding catalog data (e.g. "360�") is cleaned on output.
"""

import asyncio
import io
import json
import os
import re
import sys
import urllib.parse
import urllib.request

# Force UTF-8 on BOTH streams. The scrapers print() the ₹ symbol; on Windows'
# default cp1252 console that raises UnicodeEncodeError *inside* a scraper's
# try-block, which silently discards freshly-scraped data and made the tool fall
# back to stale catalog values (wrong price/reviews). Reconfiguring both streams
# prevents that.
sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

from scrapers.amazon import scrape_amazon, scrape_our_amazon_listing
from app import _make_keyword, _filter_competitors
from matcher import pick_best_competitor, search_query, reference_asin_for

AMAZON_CATALOG = os.path.join(os.path.dirname(__file__), "amazon_catalog.json")
PROGRESS_PATH  = os.path.join(os.path.dirname(__file__), "compare_progress.json")
OUTPUT_PATH    = os.path.join(os.path.dirname(__file__), "product_comparison_FIXED.xlsx")


# ── Text / value helpers ─────────────────────────────────────────────────────

def clean_text(s) -> str:
    """Repair mojibake from a Windows-1252 / UTF-8 mismatch at scrape time.

    The catalog stored cp1252 byte values (0x80-0xFF) directly as Unicode
    codepoints, so 360 degree became chr(0xb0) and an en-dash became chr(0x96),
    rendering as a box glyph. Decoding each such codepoint through cp1252
    restores the intended character (0xb0 -> degree, 0x96 -> en-dash, etc.).
    """
    if not isinstance(s, str):
        return s
    out = []
    for ch in s:
        o = ord(ch)
        if 0x80 <= o <= 0xFF and o != 0xA0:   # 0xA0 = NBSP, leave as-is
            try:
                out.append(bytes([o]).decode("cp1252"))
                continue
            except Exception:
                pass
        out.append(ch)
    s = "".join(out).replace(" ", " ")
    return re.sub(r"\s{2,}", " ", s).strip()


def as_float(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def as_int(v) -> int:
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def as_list(v) -> list:
    """Catalog 'keywords'/'image_urls' may be stored as a real list or a str-repr of one."""
    if isinstance(v, list):
        return v
    if isinstance(v, str) and v.strip().startswith("["):
        try:
            import ast
            parsed = ast.literal_eval(v)
            return parsed if isinstance(parsed, list) else []
        except Exception:
            return []
    return []


def clean_amazon_url(url: str) -> str:
    """Turn a sponsored/redirect Amazon URL into a clean https://www.amazon.in/dp/ASIN link."""
    if not url:
        return url
    m = re.search(r"/dp%2F([A-Z0-9]{10})", url) or re.search(r"/dp/([A-Z0-9]{10})", url)
    if not m:
        try:
            q = urllib.parse.parse_qs(urllib.parse.urlparse(url).query)
            inner = urllib.parse.unquote(q.get("url", [""])[0])
            m = re.search(r"/dp/([A-Z0-9]{10})", inner)
        except Exception:
            m = None
    if m:
        return f"https://www.amazon.in/dp/{m.group(1)}"
    return url


# ── Progress ─────────────────────────────────────────────────────────────────

def load_catalog() -> list:
    with open(AMAZON_CATALOG, "r", encoding="utf-8") as f:
        return json.load(f)


def load_progress() -> dict:
    if os.path.exists(PROGRESS_PATH):
        with open(PROGRESS_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_progress(progress: dict):
    with open(PROGRESS_PATH, "w", encoding="utf-8") as f:
        json.dump(progress, f, indent=2, ensure_ascii=False)



def pick_top_competitor(our: dict, competitors: list) -> dict | None:
    """Pick the SAME-product, best-selling, well-rated competitor.

    Delegates to the shared matcher (matcher.pick_best_competitor) so the bulk
    Excel and the dashboard analyzer agree on which competitor we benchmark
    against. _filter_competitors (own-brand/self/off-category) is applied first.
    """
    return pick_best_competitor(our, competitors, filter_fn=_filter_competitors)


# ── Build one comparison row ─────────────────────────────────────────────────

def missing_keywords(our: dict, comp: dict, limit: int = 12) -> list:
    """Keywords the competitor ranks for that OUR listing's keywords don't contain."""
    ours = {k.lower() for k in as_list(our.get("keywords"))}
    theirs = as_list(comp.get("keywords"))
    return [k for k in theirs if k.lower() not in ours][:limit]


async def compare_one(our: dict) -> dict:
    keyword = _make_keyword(our.get("title", ""))

    # Re-scrape OUR product live so price/rating/reviews/images are current — the
    # catalog values were captured earlier and drift (prices change, reviews grow).
    # Fall back to the catalog only if the live scrape genuinely returns nothing.
    live = {}
    try:
        if our.get("url"):
            live = await scrape_our_amazon_listing(our["url"]) or {}
    except Exception as e:
        print(f"   [our] live scrape failed, using catalog: {e}")

    def pick(field, default=0):
        v = live.get(field)
        return v if v not in (None, "", 0, 0.0) else default

    our_price   = as_float(pick("price",        our.get("price")))
    our_rating  = as_float(pick("rating",       our.get("rating")))
    our_reviews = as_int(  pick("review_count", our.get("review_count")))
    our_images  = as_int(  pick("image_count",  our.get("image_count")))
    our_bullets = as_int(  pick("bullet_count", our.get("bullet_count")))
    our_main_img = live.get("main_image_url") or our.get("main_image_url", "")
    our_keywords = live.get("keywords") or our.get("keywords")

    # Prefer the live title too (it's clean UTF-8 from the page, not mojibake).
    our_title = clean_text(live.get("title") or our.get("title", ""))
    our_for_kw = {"keywords": our_keywords}

    row = {
        "asin": our.get("asin", ""),
        "our_title": our_title,
        "our_url": our.get("url", ""),
        "our_img": our_main_img or (as_list(our.get("image_urls"))[:1] or [""])[0],
        "our_price": our_price, "our_rating": our_rating, "our_reviews": our_reviews,
        "our_images": our_images, "our_bullets": our_bullets,
        "keyword": keyword,
        "comp_title": "", "comp_url": "", "comp_img": "",
        "comp_price": 0, "comp_rating": 0, "comp_reviews": 0,
        "comp_images": 0, "comp_bullets": 0,
        "price_delta": 0, "cheaper": "", "rating_gap": 0, "reviews_gap": 0,
        "image_gap": 0, "bullet_gap": 0, "missing_keywords": [],
        "status": "no_competitor",
    }
    try:
        top = None
        # 0. HUMAN REFERENCE — if you marked an accurate competitor for this product,
        # scrape that exact listing directly (ground truth always wins).
        ref_asin = reference_asin_for(our.get("asin", ""))
        if ref_asin:
            ref_url = f"https://www.amazon.in/dp/{ref_asin}"
            ref = await scrape_our_amazon_listing(ref_url)
            if ref and ref.get("title"):
                top = {**ref, "url": ref_url, "source": "reference"}

        if top is None:
            # Search with the FULL descriptive title first so Amazon's own relevance
            # surfaces the exact product (e.g. 'saree covers ... clear window wardrobe'
            # rather than the lossy 5-word 'non woven foldable saree'). Fall back to the
            # short keyword only if the rich query yields no genuine same-product peer.
            # Pull 10 results so the best-selling well-rated peer isn't capped out.
            full_q = search_query(our_title)
            comps = await scrape_amazon(full_q, max_results=10) if full_q else []
            top = pick_top_competitor(our, comps)
            if top is None:
                comps = await scrape_amazon(keyword, max_results=10)
                top = pick_top_competitor(our, comps)
        if top:
            cp = as_float(top.get("price"))
            cr = as_float(top.get("rating"))
            crev = as_int(top.get("review_count"))
            cimg = as_int(top.get("image_count"))
            cbul = as_int(top.get("bullet_count"))
            row.update({
                "comp_title": clean_text(top.get("title", "")),
                "comp_url": clean_amazon_url(top.get("url", "")),
                "comp_img": top.get("main_image_url", ""),
                "comp_price": cp, "comp_rating": cr, "comp_reviews": crev,
                "comp_images": cimg, "comp_bullets": cbul,
                "price_delta": round(our_price - cp, 2) if cp else 0,
                "cheaper": ("Us" if cp and our_price < cp else
                            "Competitor" if cp and our_price > cp else
                            "Same" if cp else ""),
                "rating_gap": round(our_rating - cr, 2),
                "reviews_gap": our_reviews - crev,
                "image_gap": our_images - cimg,
                "bullet_gap": our_bullets - cbul,
                "missing_keywords": missing_keywords(our_for_kw, top),
                "status": "ok",
            })
    except Exception as e:
        row["status"] = f"error: {e}"
    return row


# ── Excel writer (with embedded thumbnails) ──────────────────────────────────

def _fetch_image(url: str):
    """Download an image URL into an openpyxl Image, scaled to ~90px. None on failure."""
    if not url:
        return None
    try:
        from openpyxl.drawing.image import Image as XLImage
        from PIL import Image as PILImage
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        data = urllib.request.urlopen(req, timeout=15).read()
        pil = PILImage.open(io.BytesIO(data)).convert("RGB")
        pil.thumbnail((90, 90))
        buf = io.BytesIO()
        pil.save(buf, format="PNG")
        buf.seek(0)
        img = XLImage(buf)
        return img
    except Exception as e:
        print(f"   [img] skip {url[:50]}: {e}")
        return None


def write_excel(rows: list):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison"

    headers = [
        "ASIN", "Our Image", "Our Product", "Our Link",
        "Comp Image", "Top Competitor", "Competitor Link",
        "Our ₹", "Comp ₹", "Price Δ", "Cheaper",
        "Our Rating", "Comp Rating", "Rating Gap",
        "Our Reviews", "Comp Reviews", "Reviews Gap",
        "Our Imgs", "Comp Imgs", "Img Gap",
        "Our Bullets", "Comp Bullets",
        "Missing Keywords (competitor has, we don't)",
        "Search Keyword", "Status",
        "the accurate link ",
    ]
    ws.append(headers)

    head_fill = PatternFill("solid", fgColor="1A1A2E")
    head_font = Font(bold=True, color="FFFFFF", size=10)
    for c in ws[1]:
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    good = Font(color="0A7D2C", bold=True)   # green – we win
    bad  = Font(color="C0392B", bold=True)   # red  – we lose

    for i, r in enumerate(rows, start=2):
        ws.append([
            r["asin"], "", r["our_title"], "Open",
            "", r["comp_title"], "Open" if r["comp_url"] else "",
            r["our_price"] or "", r["comp_price"] or "", r["price_delta"] or "", r["cheaper"],
            r["our_rating"] or "", r["comp_rating"] or "", r["rating_gap"] or "",
            r["our_reviews"] or "", r["comp_reviews"] or "", r["reviews_gap"] or "",
            r["our_images"] or "", r["comp_images"] or "", r["image_gap"] or "",
            r["our_bullets"] or "", r["comp_bullets"] or "",
            ", ".join(r["missing_keywords"]),
            r["keyword"], r["status"],
            # Pre-fill the accurate link for products you've already verified, so the
            # file always carries your confirmed answers; blank = please review.
            (f"https://www.amazon.in/dp/{reference_asin_for(r['asin'])}"
             if reference_asin_for(r["asin"]) else ""),
        ])
        ws.row_dimensions[i].height = 70

        # Hyperlinks (col D = our link 4, G = competitor link 7)
        if r["our_url"]:
            cell = ws.cell(i, 4); cell.hyperlink = r["our_url"]; cell.style = "Hyperlink"
        if r["comp_url"]:
            cell = ws.cell(i, 7); cell.hyperlink = r["comp_url"]; cell.style = "Hyperlink"

        # Color the gap columns: green when we're ahead, red when behind.
        #  Cheaper(11): green if "Us"      Rating gap(14)/Reviews gap(17)/Img gap(20): green if >0
        if r["cheaper"] == "Us":         ws.cell(i, 11).font = good
        elif r["cheaper"] == "Competitor": ws.cell(i, 11).font = bad
        for col, val in ((14, r["rating_gap"]), (17, r["reviews_gap"]), (20, r["image_gap"])):
            if val and val > 0:   ws.cell(i, col).font = good
            elif val and val < 0: ws.cell(i, col).font = bad

        # Embed thumbnails (col B = our image 2, E = comp image 5)
        our_img = _fetch_image(r.get("our_img", ""))
        if our_img:
            ws.add_image(our_img, f"B{i}")
        comp_img = _fetch_image(r.get("comp_img", ""))
        if comp_img:
            ws.add_image(comp_img, f"E{i}")

    # Borders + wrap on text columns
    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for c in row_cells:
            c.border = border
            if c.column in (3, 6, 23):
                c.alignment = Alignment(vertical="center", wrap_text=True)
            else:
                c.alignment = Alignment(vertical="center", horizontal="center")

    widths = {
        1: 13, 2: 14, 3: 42, 4: 8, 5: 14, 6: 42, 7: 10,
        8: 9, 9: 9, 10: 9, 11: 11, 12: 10, 13: 10, 14: 10,
        15: 11, 16: 12, 17: 11, 18: 8, 19: 9, 20: 8, 21: 10, 22: 11,
        23: 50, 24: 24, 25: 14, 26: 45,
    }
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w

    # Tint the "accurate link" review column (26) light yellow. Rows that already
    # have a verified link (your locked references) are tinted green instead.
    review_fill = PatternFill("solid", fgColor="FFF8CC")
    verified_fill = PatternFill("solid", fgColor="D6F5D6")
    for i in range(2, ws.max_row + 1):
        ws.cell(i, 26).fill = verified_fill if ws.cell(i, 26).value else review_fill

    ws.freeze_panes = "A2"

    # Save — but if the file is open in Excel (Windows locks it), fall back to a
    # numbered name instead of silently failing. This was the #1 "nothing changed"
    # trap: the writer couldn't overwrite an open workbook.
    try:
        wb.save(OUTPUT_PATH)
        print(f"Saved {OUTPUT_PATH}")
    except (PermissionError, OSError):
        base, ext = os.path.splitext(OUTPUT_PATH)
        for n in range(1, 100):
            alt = f"{base}_{n}{ext}"
            try:
                wb.save(alt)
                print(f"\n*** {OUTPUT_PATH} was OPEN/locked (close it in Excel). "
                      f"Saved to {alt} instead. ***")
                return
            except (PermissionError, OSError):
                continue
        raise


# ── Main ─────────────────────────────────────────────────────────────────────

async def main():
    catalog = load_catalog()
    products = [e for e in catalog
                if str(e.get("scraped")).lower() == "true"
                and e.get("title") and e["title"] != e.get("asin")]
    progress = load_progress()

    todo = [p for p in products if p.get("asin") not in progress]
    print(f"Total products: {len(products)} | already done: {len(progress)} | to do: {len(todo)}")

    def dump_excel():
        rows = [progress[p["asin"]] for p in products if p["asin"] in progress]
        print(f"Writing Excel ({len(rows)} rows, downloading thumbnails)...")
        write_excel(rows)

    for i, our in enumerate(todo, 1):
        row = await compare_one(our)
        progress[our["asin"]] = row
        save_progress(progress)
        comp = (row["comp_title"] or "(none)")[:45]
        print(f"[{i}/{len(todo)}] {our['asin']} -> {comp}  [{row['status']}]")

    # Write the Excel once, at the end, so we don't create _1/_2 copies while the
    # file is open in Excel during the run.
    dump_excel()
    rows = [progress[p["asin"]] for p in products if p["asin"] in progress]
    print(f"\nDone. Wrote {len(rows)} rows to {OUTPUT_PATH}")


if __name__ == "__main__":
    asyncio.run(main())
