"""
Rebuild amazon_catalog.json from the 29 PARENT products (colored rows) in
'Active Listing (1)- Parent in amazon.xlsx' -> 'main sheet', then live-scrape
each parent's Amazon page so the catalog has price/rating/reviews/images/bullets.

Child variants (uncolored rows: Pack2/Pack4/samples/color variants) are skipped —
only parent listings are analyzed.

Run:  python load_parents.py
"""

import asyncio
import json
import os
import sys

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

import openpyxl
from scrapers.amazon import scrape_our_amazon_listing

XLSX = r"C:\Users\pogot\OneDrive\Desktop\all\Active Listing (1)- Parent in amazon.xlsx"
CATALOG = os.path.join(os.path.dirname(__file__), "amazon_catalog.json")

PARENT_FILL = "FFFFC000"   # the orange/gold fill that marks most parent rows


def _is_parent_row(ws, r) -> bool:
    """A parent row is any row with a non-white solid fill in col A. Most are the
    orange RGB 'FFFFC000', but at least one parent uses a THEME color (gray/blue,
    theme=3) instead of RGB — so we accept any solid fill that isn't plain white,
    not just the one RGB value. This is why an earlier RGB-only check missed a row."""
    fill = ws.cell(r, 1).fill
    if not fill or fill.patternType != "solid":
        return False
    fg = fill.fgColor
    # Orange RGB parents
    try:
        if isinstance(fg.rgb, str) and fg.rgb == PARENT_FILL:
            return True
    except Exception:
        pass
    # Theme/indexed colored parents (any non-default theme fill) — exclude the
    # implicit white/none that openpyxl reports as theme 0 / rgb 00000000.
    rgb = fg.rgb if isinstance(getattr(fg, "rgb", None), str) else None
    if rgb in (None, "00000000", "FFFFFFFF"):
        theme = getattr(fg, "theme", None)
        # theme is an int for real theme colors; a colored (non-white) theme fill
        # marks a parent. theme 0/1 are the background/white pair — skip those.
        if isinstance(theme, int) and theme not in (0, 1):
            return True
        return False
    # Any other explicit non-white RGB solid fill counts as a parent too.
    return rgb not in (None, "00000000", "FFFFFFFF")


def extract_parents() -> list:
    wb = openpyxl.load_workbook(XLSX)   # keep styles (NOT data_only) to read fills
    ws = wb["main sheet"]
    parents = []
    for r in range(2, ws.max_row + 1):
        if not _is_parent_row(ws, r):
            continue
        asin = ws.cell(r, 2).value
        name = ws.cell(r, 3).value
        price = ws.cell(r, 5).value
        if not asin:
            continue
        parents.append({
            "asin": str(asin).strip(),
            "url": f"https://www.amazon.in/dp/{str(asin).strip()}",
            "title": (name or "").strip(),
            "price": float(price) if price else 0,
            "rating": 0,
            "review_count": 0,
            "image_count": 0,
            "image_urls": [],
            "main_image_url": "",
            "description": "",
            "bullet_count": 0,
            "keywords": [],
            "scraped": False,
        })
    return parents


async def main():
    parents = extract_parents()
    print(f"Found {len(parents)} parent products in the Excel.")

    # Write the catalog immediately (unscraped) so the dashboard already lists them.
    json.dump(parents, open(CATALOG, "w", encoding="utf-8"), indent=2, ensure_ascii=False)
    print(f"Wrote {CATALOG} with {len(parents)} parents (unscraped).")

    # Live-scrape each parent.
    for i, e in enumerate(parents, 1):
        try:
            d = await scrape_our_amazon_listing(e["url"])
        except Exception as ex:
            print(f"[{i}/{len(parents)}] {e['asin']} scrape error: {ex}")
            d = {}
        if d and d.get("title"):
            e.update({
                "title": d.get("title", e["title"]),
                "price": d.get("price", e["price"]) or e["price"],
                "rating": d.get("rating", 0),
                "review_count": d.get("review_count", 0),
                "image_count": d.get("image_count", 0),
                "image_urls": d.get("image_urls", []),
                "main_image_url": d.get("main_image_url", ""),
                "description": d.get("description", ""),
                "bullet_count": d.get("bullet_count", 0),
                "keywords": d.get("keywords", []),
                "scraped": True,
            })
            tag = "OK"
        else:
            e["scraped"] = False
            tag = "FAILED (kept Excel data)"
        # Save after every product so an interruption is safe.
        json.dump(parents, open(CATALOG, "w", encoding="utf-8"), indent=2, ensure_ascii=False)
        print(f"[{i}/{len(parents)}] {e['asin']} {tag} | "
              f"R{e['rating']} {e['review_count']}rev {e['image_count']}img | {e['title'][:45]}")

    done = sum(1 for e in parents if e["scraped"])
    print(f"\nDone. Scraped {done}/{len(parents)} parents into {CATALOG}.")


if __name__ == "__main__":
    asyncio.run(main())
