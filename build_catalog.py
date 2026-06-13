"""
One-time script: reads Active Listings Links.xlsx and builds product_catalog.json.
Run: python build_catalog.py

product_catalog.json structure:
{
  "amazon": [
    {"asin": "B0GRZZ1ZDC", "url": "https://www.amazon.in/dp/B0GRZZ1ZDC",
     "title": "", "price": 0, "rating": 0, "review_count": 0,
     "image_count": 0, "image_urls": [], "main_image_url": "",
     "description": "", "bullet_count": 0, "keywords": [],
     "scraped": false},
    ...
  ],
  "blinkit": [
    {"product_id": "790034", "url": "https://blinkit.com/prn/.../prid/790034",
     "slug": "lazer-pocket-size-sewing-kit-light-pink",
     "title": "Lazer Pocket Size Sewing Kit Light Pink",
     "price": 0, "image_count": 0, "image_urls": [],
     "main_image_url": "", "description": "", "keywords": [],
     "scraped": false},
    ...
  ]
}
"""

import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

EXCEL_PATH    = r"C:\Users\pogot\Downloads\Active Listings Links.xlsx"
AMAZON_PATH   = os.path.join(os.path.dirname(__file__), "amazon_catalog.json")
BLINKIT_PATH  = os.path.join(os.path.dirname(__file__), "blinkit_catalog.json")


def slug_to_title(slug: str) -> str:
    return slug.replace("-", " ").title()


def build_catalog():
    wb = openpyxl.load_workbook(EXCEL_PATH)

    # ── Amazon ────────────────────────────────────────────────────────────────
    ws_amazon = wb["Amazon"]
    amazon_entries = []
    seen_asins = set()
    for row in ws_amazon.iter_rows(min_row=2, values_only=True):
        asin = str(row[0]).strip() if row[0] else ""
        base = str(row[1]).strip() if row[1] else ""
        if not asin or not base or not base.startswith("http"):
            continue
        if asin in seen_asins:
            continue
        seen_asins.add(asin)
        url = base.rstrip("/") + "/" + asin
        amazon_entries.append({
            "asin":          asin,
            "url":           url,
            "title":         asin,   # placeholder — will be replaced after scraping
            "price":         0.0,
            "rating":        0.0,
            "review_count":  0,
            "image_count":   0,
            "image_urls":    [],
            "main_image_url": "",
            "description":   "",
            "bullet_count":  0,
            "keywords":      [],
            "scraped":       False,
        })

    # ── Blinkit ───────────────────────────────────────────────────────────────
    ws_blinkit = wb["Blinkit"]
    blinkit_entries = []
    seen_pids = set()
    for row in ws_blinkit.iter_rows(min_row=2, values_only=True):
        pid  = row[0]
        base = str(row[1]).strip() if row[1] else ""
        if not pid or not base or not base.startswith("http"):
            continue
        pid_str = str(int(pid))
        if pid_str in seen_pids:
            continue
        seen_pids.add(pid_str)
        url  = f"https://blinkit.com/prn/prid/{pid_str}"
        slug = "unknown"
        title = slug_to_title(slug)
        blinkit_entries.append({
            "product_id":    pid_str,
            "url":           url,
            "slug":          slug,
            "title":         title,
            "price":         0.0,
            "rating":        0.0,
            "review_count":  0,
            "image_count":   0,
            "image_urls":    [],
            "main_image_url": "",
            "description":   "",
            "bullet_count":  0,
            "keywords":      [],
            "scraped":       False,
        })

    with open(AMAZON_PATH, "w", encoding="utf-8") as f:
        json.dump(amazon_entries, f, indent=2, ensure_ascii=False)

    with open(BLINKIT_PATH, "w", encoding="utf-8") as f:
        json.dump(blinkit_entries, f, indent=2, ensure_ascii=False)

    print(f"amazon_catalog.json:  {len(amazon_entries)} entries → {AMAZON_PATH}")
    print(f"blinkit_catalog.json: {len(blinkit_entries)} entries → {BLINKIT_PATH}")
    return {"amazon": amazon_entries, "blinkit": blinkit_entries}


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    build_catalog()
